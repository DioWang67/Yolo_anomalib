import os
import pandas as pd
from datetime import datetime
import cv2
from typing import List, Dict, Any
from dataclasses import asdict, is_dataclass
from .utils import ImageUtils, DetectionResults
import numpy as np
from ultralytics.utils.plotting import colors
from core.logger import DetectionLogger
import shutil
from openpyxl import load_workbook
import time
import threading
import atexit


class ResultHandler:
    """Handle result persistence (images + Excel workbook with buffering)."""

    def __init__(self, config, base_dir: str = "Result", logger: DetectionLogger = None):
        if is_dataclass(config):
            config = asdict(config)
        self.base_dir = base_dir
        self.config = config
        self.logger = logger or DetectionLogger()
        self.colors = colors
        self.excel_path = os.path.join(self.base_dir, "results.xlsx")
        self.image_utils = ImageUtils()
        self.detection_results = DetectionResults(config)
        self.columns = [
            "時間戳記", "測試編號", "產品", "區域", "模型類型", "結果", "信心分數", "異常分數",
            "顏色檢測狀態", "顏色差異值", "錯誤訊息", "標註影像路徑", "原始影像路徑", "預處理圖像路徑",
            "異常熱圖路徑", "裁剪圖像路徑", "檢查點路徑"
        ]
        os.makedirs(self.base_dir, exist_ok=True)
        if not os.path.exists(self.excel_path):
            self._initialize_excel()

        # 建立工作簿並維持於記憶體
        self.wb = load_workbook(self.excel_path)
        self.ws = self.wb.active

        # 緩衝設定
        self.buffer: List[List[Any]] = []
        self.buffer_limit = getattr(self.config, "buffer_limit", 10)
        self.flush_interval = getattr(self.config, "flush_interval", None)
        # Thread-safety for buffer and workbook operations
        self._lock = threading.Lock()

        # Non-blocking image write queue (original/processed/crops). Keep annotated sync for GUI.
        import queue as _q
        self._img_queue: _q.Queue = _q.Queue(maxsize=1000)
        self._img_stop = False
        self._img_worker = threading.Thread(target=self._img_worker_loop, daemon=True)
        self._img_worker.start()

        if self.flush_interval:
            self._timer = threading.Timer(self.flush_interval, self._periodic_flush)
            self._timer.daemon = True
            self._timer.start()

        atexit.register(self.close)

    def _initialize_excel(self) -> None:
        """Create a fresh Excel file with headers if it does not exist."""
        df = pd.DataFrame(columns=self.columns)
        df.to_excel(self.excel_path, index=False, engine='openpyxl')

    def _read_excel(self) -> pd.DataFrame:
        """Read current Excel file or its backup, with retries."""
        backup_path = self.excel_path + ".bak"
        for attempt in range(3):
            try:
                if os.path.exists(self.excel_path):
                    return pd.read_excel(self.excel_path, engine='openpyxl')
                return pd.DataFrame(columns=self.columns)
            except Exception as e:
                self.logger.logger.error(f"讀取 Excel 時發生錯誤（第{attempt + 1}次）: {str(e)}")
                time.sleep(0.5)
                if os.path.exists(backup_path):
                    self.logger.logger.warning(f"嘗試從備份讀取 {backup_path}")
                    try:
                        return pd.read_excel(backup_path, engine='openpyxl')
                    except Exception as be:
                        self.logger.logger.error(f"讀取備份失敗: {str(be)}")
        return pd.DataFrame(columns=self.columns)

    def _get_next_test_id(self) -> int:
        """Return next test id (worksheet rows + buffered rows).

        Rows are appended to the worksheet only during flush(), so we add
        the in-memory buffer length to reflect pending rows.
        """
        # 當前緩衝中的資料列尚未寫入工作表，需加總
        return self.ws.max_row + len(self.buffer)

    def _append_to_excel(self, data: Dict) -> None:
        """Append a logical row to the in-memory buffer and flush if needed.

        Args:
            data: Mapping of header -> value. Missing keys are written as empty string.
        """
        row = []
        for col in self.columns:
            value = data.get(col, "")
            if col == "時間戳記" and isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            row.append(value)

        need_flush = False
        with self._lock:
            self.buffer.append(row)
            if len(self.buffer) >= self.buffer_limit:
                need_flush = True
        if need_flush:
            self.flush()

    def _periodic_flush(self) -> None:
        """Timer callback to flush buffered rows periodically."""
        self.flush()
        if self.flush_interval:
            self._timer = threading.Timer(self.flush_interval, self._periodic_flush)
            self._timer.daemon = True
            self._timer.start()

    def _img_worker_loop(self) -> None:
        while not self._img_stop:
            try:
                item = self._img_queue.get(timeout=0.5)
            except Exception:
                continue
            if item is None:
                self._img_queue.task_done()
                break
            try:
                path, img, params = item
                try:
                    if params:
                        cv2.imwrite(path, img, params)
                    else:
                        cv2.imwrite(path, img)
                except Exception:
                    cv2.imwrite(path, img)
            except Exception:
                pass
            finally:
                self._img_queue.task_done()

    def flush(self) -> None:
        """Write buffered rows to Excel file, with backup and retry."""
        if not self.buffer:
            return
        backup_path = self.excel_path + ".bak"
        for attempt in range(3):
            try:
                if os.path.exists(self.excel_path):
                    shutil.copy(self.excel_path, backup_path)
                with self._lock:
                    for row in self.buffer:
                        self.ws.append(row)
                    self.wb.save(self.excel_path)
                    self.buffer.clear()
                if os.path.exists(backup_path):
                    os.remove(backup_path)
                self.buffer.clear()
                self.logger.logger.info(f"Excel 數據已保存至 {self.excel_path}")
                return
            except PermissionError:
                self.logger.logger.error(f"權限拒絕，無法寫入 {self.excel_path}，請檢查檔案是否被占用或權限設置")
            except Exception as e:
                self.logger.logger.error(f"寫入 Excel 時發生錯誤（第{attempt + 1}次）: {str(e)}")
            time.sleep(0.5)
        if os.path.exists(backup_path):
            shutil.copy(backup_path, self.excel_path)
            self.logger.logger.warning(f"已從備份恢復 {self.excel_path}")
            os.remove(backup_path)

    def close(self) -> None:
        """Flush, stop timer and close the workbook on exit."""
        self.flush()
        # Stop image worker
        try:
            self._img_stop = True
            try:
                self._img_queue.put_nowait(None)
            except Exception:
                pass
            try:
                self._img_queue.join()
            except Exception:
                pass
        except Exception:
            pass
        if self.flush_interval and hasattr(self, "_timer"):
            self._timer.cancel()
            if hasattr(self, "wb"):
                self.wb.close()

    def _draw_detection_box(self, frame: np.ndarray, detection: Dict) -> None:
        """Draw one detection (bbox + class/conf label) on frame."""
        x1, y1, x2, y2 = detection['bbox']
        label = f"{detection['class']} {detection['confidence']:.2f}"
        color = self.colors(detection['class_id'], True)
        cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
        self.image_utils.draw_label(frame, label, (x1, y1 - 10), color)

    def get_annotated_path(self, status: str, detector: str, product: str, area: str, anomaly_score: float = None) -> str:
        """生成標註影像的完整路徑"""
        current_date = datetime.now().strftime("%Y%m%d")
        time_stamp = datetime.now().strftime("%H%M%S")
        base_path = os.path.join(self.base_dir, current_date,product ,area, status, "annotated", detector.lower())

        if detector.lower() == "anomalib" and anomaly_score is not None:
            image_name = f"{detector.lower()}_{product}_{area}_{time_stamp}_{anomaly_score:.4f}.jpg"
        else:
            image_name = f"{detector.lower()}_{product}_{area}_{time_stamp}.jpg" if product and area else f"{detector.lower()}_{time_stamp}.jpg"

        os.makedirs(base_path, exist_ok=True)
        return os.path.join(base_path, image_name)

    def save_results(
        self,
        frame: np.ndarray,
        detections: List[Dict[str, Any]],
        status: str,
        detector: str,
        missing_items: List[str],
        processed_image: np.ndarray,
        anomaly_score: float = None,
        heatmap_path: str = None,
        product: str = None,
        area: str = None,
        ckpt_path: str = None,
        color_result: Dict[str, Any] = None,
    ) -> Dict:
        try:
            current_date = datetime.now().strftime("%Y%m%d")
            time_stamp = datetime.now().strftime("%H%M%S")
            base_path = os.path.join(self.base_dir, current_date ,product ,area , status)
            detector_prefix = detector.lower()

            os.makedirs(os.path.join(base_path, "original", detector_prefix), exist_ok=True)
            os.makedirs(os.path.join(base_path, "preprocessed", detector_prefix), exist_ok=True)
            os.makedirs(os.path.join(base_path, "annotated", detector_prefix), exist_ok=True)
            os.makedirs(os.path.join(base_path, "cropped", detector_prefix), exist_ok=True)

            image_name = f"{detector_prefix}_{product}_{area}_{time_stamp}.jpg" if product and area else f"{detector_prefix}_{time_stamp}.jpg"
            if detector_prefix == "anomalib" and anomaly_score is not None:
                image_name = f"{detector_prefix}_{product}_{area}_{time_stamp}_{anomaly_score:.4f}.jpg"

            # Saving controls
            cfg = self.config
            _get = (cfg.get if isinstance(cfg, dict) else lambda k, d=None: getattr(cfg, k, d))
            only_fail = bool(_get("save_fail_only", False))
            should_save_images = (status != "PASS") if only_fail else True
            save_original = bool(_get("save_original", True)) and should_save_images
            save_processed = bool(_get("save_processed", True)) and should_save_images
            save_annotated = bool(_get("save_annotated", True)) and should_save_images
            save_crops = bool(_get("save_crops", True)) and should_save_images
            jpeg_quality = int(_get("jpeg_quality", 95))
            png_compression = int(_get("png_compression", 3))

            imwrite_params_jpg = [int(cv2.IMWRITE_JPEG_QUALITY), int(max(1, min(100, jpeg_quality)))]
            imwrite_params_png = [int(cv2.IMWRITE_PNG_COMPRESSION), int(max(0, min(9, png_compression)))]

            original_path = os.path.join(base_path, "original", detector_prefix, image_name)
            preprocessed_path = os.path.join(base_path, "preprocessed", detector_prefix, image_name)

            # write images (async for non-annotated)
            def _enqueue_write(path, img):
                params = imwrite_params_png if path.lower().endswith('.png') else imwrite_params_jpg
                try:
                    self._img_queue.put_nowait((path, img, params))
                except Exception:
                    try:
                        cv2.imwrite(path, img, params)
                    except Exception:
                        cv2.imwrite(path, img)

            def _imwrite_sync(path, img):
                try:
                    if path.lower().endswith('.png'):
                        cv2.imwrite(path, img, imwrite_params_png)
                    else:
                        cv2.imwrite(path, img, imwrite_params_jpg)
                except Exception:
                    cv2.imwrite(path, img)

            if save_original:
                _enqueue_write(original_path, frame)
            else:
                original_path = ""

            if save_processed:
                _enqueue_write(preprocessed_path, processed_image)
            else:
                preprocessed_path = ""

            # 處理標註影像
            if detector.lower() == "yolo" and save_annotated:
                annotated_frame = processed_image.copy()
                crop_source = processed_image

                if detections:
                    # Draw per-detection boxes and color status (OK/NG) when available
                    color_items = []
                    try:
                        color_items = (color_result or {}).get("items", []) if color_result else []
                    except Exception:
                        color_items = []
                    fail_indices = []
                    for idx, det in enumerate(detections):
                        self._draw_detection_box(annotated_frame, det)
                        try:
                            if idx < len(color_items):
                                it = color_items[idx]
                                is_ok = bool(it.get("is_ok", False))
                                tag = "OK" if is_ok else "NG"
                                tag_color = (0, 200, 0) if is_ok else (0, 0, 255)
                                x1, y1, x2, y2 = det['bbox']
                                # place small tag inside box corner
                                pos = (max(0, x1 + 5), max(0, y1 + 20))
                                self.image_utils.draw_label(annotated_frame, f"{tag}", pos, tag_color, font_scale=0.7, thickness=2)
                                if not is_ok:
                                    fail_indices.append(idx)
                        except Exception:
                            pass

                status_color = (0, 255, 0) if status == "PASS" else (0, 0, 255)
                text_x = 50
                text_y = min(50, annotated_frame.shape[0] - 20)
                self.image_utils.draw_label(annotated_frame, f"Status: {status}", (text_x, text_y), status_color, font_scale=1.0, thickness=2)
                if color_result:
                    color_status = "PASS" if color_result.get("is_ok", False) else "FAIL"
                    color_color = (0, 255, 0) if color_result.get("is_ok", False) else (0, 0, 255)
                    text_y += 30
                    # Show only PASS/FAIL for color result (no diffs)
                    self.image_utils.draw_label(
                        annotated_frame,
                        f"Color: {color_status}",
                        (text_x, text_y),
                        color_color,
                        font_scale=1.0,
                        thickness=2,
                    )
                    # If any failed indices, show them for quick localization
                    try:
                        items = color_result.get("items", [])
                        bad = [str(i) for i, it in enumerate(items) if not it.get("is_ok", False)]
                        if bad:
                            text_y += 30
                            self.image_utils.draw_label(
                                annotated_frame,
                                f"NG idx: {', '.join(bad)}",
                                (text_x, text_y),
                                (0, 0, 255),
                                font_scale=1.0,
                                thickness=2,
                            )
                    except Exception:
                        pass

                annotated_path = os.path.join(base_path, "annotated", detector_prefix, image_name)
                _imwrite_sync(annotated_path, annotated_frame)

            else:  # Anomalib or no annotated
                annotated_path = os.path.join(base_path, "annotated", detector_prefix, image_name) if save_annotated else ""
                crop_source = frame

            # 處理裁剪圖像（僅對 YOLO 有效）
            cropped_paths = []
            if detector.lower() == "yolo" and detections and save_crops:
                max_crops = _get("max_crops_per_frame", None)
                for idx, det in enumerate(detections):
                    if max_crops is not None and idx >= int(max_crops):
                        break
                    x1, y1, x2, y2 = det['bbox']
                    x1, y1 = max(0, x1), max(0, y1)
                    x2, y2 = min(crop_source.shape[1], x2), min(crop_source.shape[0], y2)
                    cropped_img = crop_source[y1:y2, x1:x2]
                    cropped_filename = f"{detector_prefix}_{product}_{area}_{time_stamp}_{det['class']}_{idx}.png"
                    cropped_path = os.path.join(base_path, "cropped", detector_prefix, cropped_filename)
                    _enqueue_write(cropped_path, cropped_img)
                    cropped_paths.append(cropped_path)

            # 處理熱圖路徑（僅對 Anomalib 有效）
            heatmap_dest_path = annotated_path if (save_annotated and detector.lower() == "anomalib" and heatmap_path and os.path.exists(heatmap_path)) else None

            confidence_scores = ";".join([f"{det['class']}:{det['confidence']:.2f}" for det in detections]) if detections else ""
            error_message = "" if status == "PASS" else f"缺少元件: {', '.join(missing_items)}" if missing_items else "異常分數超出閾值"
            color_status = ""
            diff_value = ""
            if color_result:
                color_status = "PASS" if color_result.get("is_ok", False) else "FAIL"
                diff_value = ";".join([
                    f"{item.get('diff', 0):.2f}" for item in color_result.get("items", [])
                ])

            excel_data = {
                "時間戳記": datetime.now(),
                "測試編號": self._get_next_test_id(),
                "產品": product,
                "區域": area,
                "模型類型": detector,
                "結果": status,
                "信心分數": confidence_scores,
                "異常分數": anomaly_score,
                "顏色檢測狀態": color_status,
                "顏色差異值": diff_value,
                "錯誤訊息": error_message,
                "標註影像路徑": annotated_path,
                "原始影像路徑": original_path,
                "預處理圖像路徑": preprocessed_path,
                "異常熱圖路徑": heatmap_dest_path if heatmap_dest_path else "",
                "裁剪圖像路徑": ";".join(cropped_paths) if cropped_paths else "",
                "檢查點路徑": ckpt_path or "",
            }
            self._append_to_excel(excel_data)

            return {
                "status": "SUCCESS",
                "original_path": original_path,
                "preprocessed_path": preprocessed_path,
                "annotated_path": annotated_path,
                "heatmap_path": heatmap_dest_path,
                "cropped_paths": cropped_paths,
                "product": product,
                "area": area,
            }
        except Exception as e:
            self.logger.logger.error(f"保存結果失敗: {str(e)}")
            return {"status": "ERROR", "error_message": str(e)}

