from __future__ import annotations

"""具備備份與批次寫入的 Excel 緩衝寫手，降低寫入失敗風險。"""

import os
import shutil
import threading
import time
from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

import pandas as pd  # type: ignore[import]
from openpyxl import load_workbook  # type: ignore[import]


@dataclass
class ExcelFlushResult:
    success: bool
    rows_written: int = 0
    error: str | None = None


@dataclass
class ExcelWorkbookBuffer:
    """Buffered Excel writer with backups and optional periodic flush."""

    path: str
    columns: list[str]
    buffer_limit: int
    logger: Any
    flush_interval: float | None = None
    workbook_kwargs: dict[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        self.buffer: list[list[Any]] = []
        self._lock = threading.Lock()
        self.backup_path = self.path + ".bak"
        os.makedirs(os.path.dirname(self.path), exist_ok=True)
        if not os.path.exists(self.path):
            self._initialize_excel()
        self.wb = load_workbook(self.path, **self.workbook_kwargs)
        self.ws = self.wb.active
        self._timer: threading.Timer | None = None
        if self.flush_interval:
            self._timer = threading.Timer(
                self.flush_interval, self._periodic_flush
            )
            self._timer.daemon = True
            self._timer.start()

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def append(self, row: Iterable[Any]) -> None:
        row_list = list(row)
        with self._lock:
            self.buffer.append(row_list)
            should_flush = len(self.buffer) >= max(1, self.buffer_limit)
        if should_flush:
            self.flush()

    def flush(self) -> ExcelFlushResult:
        with self._lock:
            if not self.buffer:
                return ExcelFlushResult(success=True, rows_written=0)
            rows = list(self.buffer)
            self.buffer.clear()
        for attempt in range(3):
            try:
                if os.path.exists(self.path):
                    shutil.copy(self.path, self.backup_path)
                for row in rows:
                    self.ws.append(row)
                self.wb.save(self.path)
                if os.path.exists(self.backup_path):
                    os.remove(self.backup_path)
                self.logger.info(f"Excel 已更新: {self.path}")
                return ExcelFlushResult(success=True, rows_written=len(rows))
            except PermissionError:
                self.logger.error(

                        f"權限不足，無法寫入 {self.path}，"
                        "請檢查檔案是否開啟或權限設定"

                )
            except Exception as exc:
                self.logger.error(f"寫入 Excel 發生錯誤 (第{attempt + 1}次重試): {exc}")
            time.sleep(0.5)
        if os.path.exists(self.backup_path):
            shutil.copy(self.backup_path, self.path)
            self.logger.warning(
                f"已從備份還原 Excel: {self.path}"
            )
            os.remove(self.backup_path)
        return ExcelFlushResult(
            success=False,
            rows_written=0,
            error="flush_failed"
        )

    def next_test_id(self, pending_count: int = 0) -> int:
        return self.ws.max_row + pending_count

    def pending_rows(self) -> int:
        with self._lock:
            return len(self.buffer)

    def close(self) -> None:
        if self._timer:
            self._timer.cancel()
        self.flush()
        self.wb.close()

    # ------------------------------------------------------------------
    # Internals
    # ------------------------------------------------------------------

    def _initialize_excel(self) -> None:
        df = pd.DataFrame(columns=self.columns)
        df.to_excel(self.path, index=False, engine="openpyxl")

    def _periodic_flush(self) -> None:
        try:
            self.flush()
        finally:
            if self.flush_interval:
                self._timer = threading.Timer(
                    self.flush_interval, self._periodic_flush
                )
                self._timer.daemon = True
                self._timer.start()


def format_excel_row(columns: list[str], data: dict[str, Any]) -> list[Any]:
    row: list[Any] = []
    for col in columns:
        value = data.get(col, "")
        if isinstance(value, datetime):
            value = value.strftime("%Y-%m-%d %H:%M:%S")
        row.append(value)
    return row
