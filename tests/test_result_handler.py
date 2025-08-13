import os
from datetime import datetime
from openpyxl import load_workbook

from core.result_handler import ResultHandler


def dummy_config():
    return {}


def test_append_to_excel_multiple_times(tmp_path):
    handler = ResultHandler(dummy_config(), base_dir=str(tmp_path))
    base_data = {
        "時間戳記": datetime.now(),
        "產品": "prod",
        "區域": "A",
        "模型類型": "yolo",
        "結果": "PASS",
        "信心分數": "cls:0.9",
        "異常分數": 0.1,
        "錯誤訊息": "",
        "標註影像路徑": "p1",
        "原始影像路徑": "p2",
        "預處理圖像路徑": "p3",
        "異常熱圖路徑": "",
        "裁剪圖像路徑": "",
        "檢查點路徑": "",
    }

    for i in range(5):
        data = dict(base_data)
        data["測試編號"] = i + 1
        handler._append_to_excel(data)

    wb = load_workbook(handler.excel_path)
    sheet = wb.active
    assert sheet.max_row == 6  # 1 header + 5 data rows
    ids = [row[1].value for row in sheet.iter_rows(min_row=2, max_row=6, values_only=True)]
    assert ids == [1, 2, 3, 4, 5]
